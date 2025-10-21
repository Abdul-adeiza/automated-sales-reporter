# Import all required libraries
from pathlib import Path
import openpyxl
from openpyxl.styles import numbers, Font, PatternFill, Alignment, Border, Side
import re
import csv
import requests
import time
from datetime import datetime
import shutil


# Get required folder for processing
BASE_DIR = Path(__file__).resolve().parent
INCOMING_FOLDER = BASE_DIR / "incoming_orders"
ARCHIVE_FOLDER = BASE_DIR / "archive"
REPORT_FILE = BASE_DIR / "processed orders summary.xlsx"


# Define regex pattern for information filtering
order_id_regex = r"(?:OrderID|Order|OID|Transaction ID|Order Number|RefNo|id|order_ref|order|transaction)[#:=\s>\-\"{]*([A-Z]-?\d{3,})"
product_id_regex = r"(?:ProductID|Product|PID|Item Code|SKU|prod_id|product_id|product|P_ID)[#:=\s>\-/\"{]*([0-9]+)"
quantity_regex = r"(?:Quantity|Qty|qty|Amount|Units|count|Q|quantity)[#:=\s>\-/\"{]*([0-9]+)"

# Define a function to read data from a text file and apply regular expression
def read_text_file(file_path, order_id_rgx, product_id_rgx, quantity_rgx):
    """
    Reads a text file and extracts order details using regex.
    Returns a dictionary with the details if all three are found, otherwise returns None.
    """
    with open(file_path, 'r') as file:
        content = file.read()

    order_id_match = re.search(order_id_rgx, content)
    product_id_match = re.search(product_id_rgx, content)
    quantity_match = re.search(quantity_rgx, content)

    # logic: if all three matches are found, return the data
    if order_id_match and product_id_match and quantity_match:
        return {
            "order_id": order_id_match.group(1),
            "product_id": product_id_match.group(1),
            "quantity": quantity_match.group(1)
        }

    print(f"Warning: Skipping {file_path.name} due to missing data.")
    return None

def main():
    # Filter and collect data from text and csv file into a list of dictionaries
    parsed_orders = []
    for item in INCOMING_FOLDER.iterdir():
        if item.is_file() and item.suffix == ".txt":
            read_data = read_text_file(item, order_id_regex, product_id_regex, quantity_regex)
            if read_data:  # Only add the data if function doesn't return None
                parsed_orders.append(read_data)

        elif item.is_file() and item.suffix == ".csv":
            with open(item, "r") as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=",")
                next(csv_reader)  # Skip header row
                for row in csv_reader:
                    # Convert the CSV row into the same dictionary structure
                    parsed_orders.append({
                        "order_id": row[0],
                        "product_id": row[1],
                        "quantity": row[2]
                    })


    # Get a unique set of product IDs to avoid redundant API calls
    unique_product_ids = {order['product_id'] for order in parsed_orders}

    web_detail_dict = {}
    for product_id in unique_product_ids:
        url = f'https://fakestoreapi.com/products/{product_id}'
        response = requests.get(url, headers={"Accept": "application/json"})

        if response.status_code == 429:
            print("Rate limit hit. Waiting before retrying...")
            time.sleep(60)
        elif response.ok and response.text.strip():
            try:
                data = response.json()
                web_detail_dict[product_id] = data
            except ValueError:
                print(f"Response for Product ID {product_id} is not valid JSON.")
        else:
            print(f"Empty or failed response for Product ID {product_id}. Status code: {response.status_code}")


    # Compile all products detail into a list of list for Excel
    final_product_details = [['Order_ID', 'ProductID', 'Name', 'Category', 'Quantity', 'Price ($)', 'Rating', 'Count', 'Total Price ($)']]

    for order in parsed_orders:
        product_id = order['product_id']
        if product_id in web_detail_dict:
            product_info = web_detail_dict[product_id]

            # Calculate total price here in Python
            try:
                quantity = int(order['quantity'])
                price = float(product_info['price'])
                total_price = quantity * price
            except (ValueError, TypeError):
                total_price = 0.0

            inner_list = [
                order['order_id'],
                product_id,
                product_info['title'],
                product_info['category'],
                quantity,
                price,
                product_info['rating']['rate'],
                product_info['rating']['count'],
                total_price  # Add the calculated total price
            ]
            final_product_details.append(inner_list)


    # Write file list into an Excel file
    # Check if the report file exists to avoid an error on the first run
    if REPORT_FILE.exists():
        workbook = openpyxl.load_workbook(REPORT_FILE)
    else:
        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

    date = datetime.today()
    sheet_name = date.strftime('%Y-%m-%d')
    daily_orders = workbook.create_sheet(sheet_name)

    # Create the serial number columns
    daily_orders['A1'] = 'S/N'
    daily_orders['A2'] = '1'
    sn_address = f'A3:A{str(len(final_product_details))}'
    sn = daily_orders[sn_address]
    for idx, cell in enumerate(sn):
        current_row = idx + 3
        formula_cell = 'A' + str(current_row)
        formula = f'=A{current_row - 1} + 1'
        daily_orders[formula_cell] = formula

    range_address = 'B1:J' + str(len(final_product_details))
    rng = daily_orders[range_address]
    for i in range(len(final_product_details)):
        row = final_product_details[i]
        for j in range(len(row)):
            val = row[j]
            rng[i][j].value = val


    # Create Dashboard sheet
    dash_board = workbook.create_sheet('Dashboard')

    # Call the daily transaction sheet
    last_row = daily_orders.max_row

    # Populate the Dashboard sheet
    dash_board['E1'] = 'Daily Sales Dashboard'
    dash_board['F1'] = f"Report For: {sheet_name}"
    dash_board['E2'] = 'Key Metrics'

    dash_board['E3'] = 'Total Revenue'
    dash_board['F3'] = f"=SUM('{sheet_name}'!J2:J{last_row})"

    dash_board['E4'] = 'Total Item Sold'
    dash_board['F4'] = f"=SUMPRODUCT(--('{sheet_name}'!F2:F{last_row}))"

    dash_board['E5'] = 'Number of Orders'
    dash_board['F5'] = f"=COUNTA('{sheet_name}'!B2:B{last_row})"

    dash_board['E6'] = 'Sales by Category'

    dash_board['E7'] = "Mens's Clothing"
    category_men = "men's clothing"
    dash_board['F7'] = f"=SUMIF('{sheet_name}'!E2:E{last_row}, \"{category_men}\", '{sheet_name}'!J2:J{last_row})"

    dash_board['E8'] = "Women's Clothing"
    category_women = "women's clothing"
    dash_board['F8'] = f"=SUMIF('{sheet_name}'!E2:E{last_row}, \"{category_women}\", '{sheet_name}'!J2:J{last_row})"

    dash_board['E9'] = "Electronics"
    category_electronics = "electronics"
    dash_board['F9'] = f"=SUMIF('{sheet_name}'!E2:E{last_row}, \"{category_electronics}\", '{sheet_name}'!J2:J{last_row})"

    dash_board['E10'] = "Jewelery"
    category_jewelery = "jewelery"
    dash_board['F10'] = f"=SUMIF('{sheet_name}'!E2:E{last_row}, \"{category_jewelery}\", '{sheet_name}'!J2:J{last_row})"


    # Format daily_orders and dash_board Sheet
    # Format All Sum Values as Currency
    currency_cells = ['F3', 'F7', 'F8', 'F9', 'F10']
    for cell in currency_cells:
        dash_board[cell].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Bold and Patter Fill Key Labels
    bold_font = Font(bold=True)
    highlight_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    highlight_cells = ['E1', 'F1', 'E2', 'E6']
    for cell in highlight_cells:
        dash_board[cell].font = bold_font
        dash_board[cell].fill = highlight_fill
    header_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    for cell in daily_orders[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # Pattern Fill and Bold for the Price Columns (F=Quantity, G=Price, J=Total Price)
    price_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    for row in daily_orders.iter_rows(min_row=2, max_row=daily_orders.max_row):
        for col in [6, 7, 10]:
            cell = row[col - 1]
            cell.fill = price_fill
            cell.font = bold_font

    # Merge and Center Section Titles (Dashboard)
    dash_board.merge_cells('E2:F2')
    dash_board.merge_cells('E6:F6')
    center_align = Alignment(horizontal='center', vertical='center')
    dash_board['E2'].alignment = center_align
    dash_board['E6'].alignment = center_align

    # Apply Thick Boarder Around Dashboard
    thick_side = Side(border_style='thick', color='000000')
    thick_border = Border(top=thick_side, bottom=thick_side, left=thick_side, right=thick_side)
    for row in daily_orders.iter_rows(min_row=1, max_row=daily_orders.max_row, min_col=1, max_col=daily_orders.max_column):
        for cell in row:
            cell.border = thick_border
    for row in dash_board.iter_rows(min_row=1, max_row=10, min_col=5, max_col=6):
        for cell in row:
            cell.border = thick_border

    # Fill Header Row with color
    header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type='solid')
    dash_board['E1'].fill = header_fill
    dash_board['F1'].fill = header_fill

    # Save changes to workbook
    workbook.save(REPORT_FILE)

    # Move processed files form INCOMING_FOLDER to ARCHIVE_FOLDER
    archive_subfolder = ARCHIVE_FOLDER / sheet_name
    archive_subfolder.mkdir(parents=True, exist_ok=True)

    for src_path in INCOMING_FOLDER.iterdir():
        if src_path.is_file():
            dest_path = archive_subfolder / src_path.name

            if dest_path.exists():
                stem = src_path.stem
                suffix = src_path.suffix
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                dest_path = archive_subfolder / f"{stem}_{timestamp}{suffix}"

            try:
                shutil.move(str(src_path), str(dest_path))
            except Exception as e:
                print(f"Failed to move {src_path.name} to archive {e}")

if __name__ == "__main__":
    main()

